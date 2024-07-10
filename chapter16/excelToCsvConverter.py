import openpyxl
import csv
import os

# Load the Excel workbook
excelFile = 'example.xlsx'  # Replace with your actual Excel file name
workbook = openpyxl.load_workbook(excelFile)

# Create a directory to store the CSV files
os.makedirs('csv_files', exist_ok=True)

# Loop through each sheet in the workbook
for sheetName in workbook.sheetnames:
    sheet = workbook[sheetName]
    
    # Create a CSV file with the same name as the sheet
    csvFilename = f'csv_files/{sheetName}.csv'
    with open(csvFilename, 'w', newline='') as csvFile:
        csvWriter = csv.writer(csvFile)
        
        # Loop through each row in the sheet
        for row in sheet.iter_rows(values_only=True):
            csvWriter.writerow(row)

print('Conversion complete! CSV files are saved in the "csv_files" directory.')
