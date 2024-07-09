import openpyxl
from openpyxl.utils import get_column_letter

def create_multiplication_table(size):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Populate the first row and first column
    for i in range(1, size + 1):
        sheet.cell(row=1, column=i+1).value = i
        sheet.cell(row=i+1, column=1).value = i

    # Fill in the multiplication table
    for row in range(2, size + 2):
        for col in range(2, size + 2):
            sheet.cell(row=row, column=col).value = (row - 1) * (col - 1)

    # Save the workbook
    workbook.save('multiplication_table.xlsx')

# Define the size of the multiplication table
table_size = 10  # You can change this value to create a larger or smaller table

# Create the multiplication table
create_multiplication_table(table_size)
