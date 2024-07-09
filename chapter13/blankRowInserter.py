import openpyxl

def insert_blank_rows(filename, start_row, num_rows):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Get the maximum row and column in the sheet
    max_row = sheet.max_row
    max_col = sheet.max_column

    # Shift the rows down
    for row in range(max_row, start_row - 1, -1):
        for col in range(1, max_col + 1):
            new_cell = sheet.cell(row=row + num_rows, column=col)
            old_cell = sheet.cell(row=row, column=col)
            new_cell.value = old_cell.value
            old_cell.value = None

    # Save the workbook
    new_filename = 'blank_row_inserted_' + filename
    workbook.save(new_filename)

# Define the parameters
filename = 'example.xlsx'  # The Excel file to modify
start_row = 3              # The row to start inserting blank rows
num_rows = 2               # The number of blank rows to insert

# Insert blank rows
insert_blank_rows(filename, start_row, num_rows)
