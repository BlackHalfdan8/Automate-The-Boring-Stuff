import openpyxl
import os

def spreadsheet_to_text_files(filename, output_directory):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Make sure the output directory exists
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    # Get the maximum row and column in the sheet
    max_row = sheet.max_row
    max_col = sheet.max_column

    # Loop through each column
    for col_num in range(1, max_col + 1):
        # Create a text file for each column
        with open(os.path.join(output_directory, f'column_{col_num}.txt'), 'w') as file:
            # Write each cell in the column to the text file
            for row_num in range(1, max_row + 1):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    file.write(str(cell_value) + '\n')

# Define the parameters
filename = 'example.xlsx'           # The Excel file to read from
output_directory = 'text_files'     # The directory to save text files

# Convert spreadsheet to text files
spreadsheet_to_text_files(filename, output_directory)
