import openpyxl

def invert_spreadsheet(filename):
    # Load the original workbook and select the active sheet
    original_wb = openpyxl.load_workbook(filename)
    original_sheet = original_wb.active

    # Create a new workbook and select the active sheet
    inverted_wb = openpyxl.Workbook()
    inverted_sheet = inverted_wb.active

    # Get the maximum row and column in the original sheet
    max_row = original_sheet.max_row
    max_col = original_sheet.max_column

    # Invert the cells
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            original_cell_value = original_sheet.cell(row=row, column=col).value
            inverted_sheet.cell(row=col, column=row).value = original_cell_value

    # Save the new workbook
    new_filename = 'inverted_' + filename
    inverted_wb.save(new_filename)

# Define the parameters
filename = 'example.xlsx'  # The Excel file to invert

# Invert the spreadsheet
invert_spreadsheet(filename)
